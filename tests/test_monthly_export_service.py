from datetime import datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.monthly_export_service import MonthlyExportService


class ReportRepositoryStub:
    def list_purchases_for_month(self, **kwargs):
        person = SimpleNamespace(name="Tomas")
        relation = SimpleNamespace(
            person=person,
            shared_value=Decimal("40.00"),
        )
        return [
            SimpleNamespace(
                purchase_value=Decimal("120.00"),
                purchase_date=datetime(2026, 7, 10, 12, 0),
                purchase_place="Restaurante",
                notes="Almoco",
                is_installment=False,
                installments=[],
                is_shared=True,
                people=[relation],
            )
        ]


class ReceivableRepositoryStub:
    def list_open_summary_for_month(self, **kwargs):
        return [
            SimpleNamespace(
                person_name="Tomas",
                total=Decimal("40.00"),
                pending_count=1,
            )
        ]

    def list_open_details_for_month(self, **kwargs):
        return [
            SimpleNamespace(
                person_name="Tomas",
                amount=Decimal("40.00"),
                purchase_date=datetime(2026, 7, 10, 12, 0),
                purchase_place="Restaurante",
                notes="Almoco",
            )
        ]


def test_builds_monthly_workbook_in_requested_layout():
    service = MonthlyExportService(
        report_repository=ReportRepositoryStub(),
        receivable_repository=ReceivableRepositoryStub(),
    )

    content, filename = service.build(
        year=2026,
        month=7,
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Relatorio mensal"]

    assert filename == "financebot_relatorio_2026_07.xlsx"
    assert [sheet.cell(1, column).value for column in range(1, 8)] == [
        "Valor",
        "Data",
        "Obs",
        "Parcelado",
        "Nr parcelas",
        "Dividido",
        "Pessoas divididas",
    ]
    assert sheet["I1"].value == "Quem deve?"
    assert sheet["J1"].value == "Quanto deve"
    assert sheet["I2"].value == "Tomas"
    assert sheet["J2"].value == 40
    assert "Detalhes por pessoa" in workbook.sheetnames
