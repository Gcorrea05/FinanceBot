from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path
import re
import unicodedata


MAX_IMPORT_ROWS = 5000
MAX_INSPECTION_ROWS = 30


class ImportFileError(ValueError):
    pass


class IgnoredImportRow(ValueError):
    pass


@dataclass(frozen=True)
class ImportColumnMapping:
    data_start_row: int
    date_column: int
    description_columns: tuple[int, ...]
    amount_column: int
    external_id_column: int | None = None
    sheet_name: str | None = None
    date_format: str = "auto"
    decimal_separator: str = "auto"
    amount_mode: str = "all"


@dataclass(frozen=True)
class ImportInspection:
    source_type: str
    sheets: tuple[str, ...]
    selected_sheet: str | None
    total_rows: int
    max_columns: int
    rows: tuple[tuple[str, ...], ...]
    mapping_required: bool


@dataclass(frozen=True)
class ParsedImportRow:
    row_number: int
    purchase_date: datetime | None
    purchase_place: str | None
    purchase_value: Decimal | None
    external_id: str | None
    fingerprint: str | None
    classification: str = "valid"
    error_message: str | None = None

    @property
    def valid(self) -> bool:
        return self.classification == "valid"

    @property
    def ignored(self) -> bool:
        return self.classification == "ignored"


@dataclass(frozen=True)
class ParsedImportFile:
    source_type: str
    rows: tuple[ParsedImportRow, ...]


def _normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _clean_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return " ".join(str(value or "").strip().split())


def _trim_row(values: list[object]) -> list[object]:
    trimmed = list(values)
    while trimmed and not _clean_text(trimmed[-1]):
        trimmed.pop()
    return trimmed


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportFileError("Nao foi possivel identificar a codificacao do CSV.")


def _read_csv_rows(content: bytes) -> list[list[object]]:
    text = _decode_csv(content)
    sample_lines = [line for line in text.splitlines()[:20] if line.strip()]
    first_line = sample_lines[0] if sample_lines else ""
    candidates = (";", "\t", "|", ",")
    delimiter = max(candidates, key=lambda candidate: first_line.count(candidate))

    if first_line.count(delimiter) == 0:
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ";"

    reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows = [_trim_row(list(row)) for row in reader]
    return [row for row in rows if any(_clean_text(value) for value in row)]


def _load_xlsx(content: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ImportFileError("A dependencia openpyxl nao esta instalada.") from error

    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ImportFileError("O arquivo XLSX nao pode ser lido.") from error


def _read_xlsx_rows(content: bytes, sheet_name: str | None) -> tuple[list[list[object]], tuple[str, ...], str]:
    workbook = _load_xlsx(content)
    sheets = tuple(workbook.sheetnames)
    if not sheets:
        raise ImportFileError("A planilha nao possui abas.")

    selected = sheet_name or sheets[0]
    if selected not in sheets:
        raise ImportFileError(f"A aba '{selected}' nao existe no arquivo.")

    sheet = workbook[selected]
    rows = [_trim_row(list(values)) for values in sheet.iter_rows(values_only=True)]
    rows = [row for row in rows if any(_clean_text(value) for value in row)]
    return rows, sheets, selected


def _ofx_value(block: str, tag: str) -> str:
    match = re.search(rf"<{tag}>([^<\r\n]+)", block, flags=re.IGNORECASE)
    return _clean_text(match.group(1)) if match else ""


def _ofx_blocks(content: bytes) -> list[str]:
    text = _decode_csv(content)
    blocks = re.findall(
        r"<STMTTRN>(.*?)(?:</STMTTRN>|(?=<STMTTRN>)|$)",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not blocks:
        raise ImportFileError("Nenhuma transacao STMTTRN foi encontrada no OFX.")
    return blocks


def _read_tabular_file(
    filename: str,
    content: bytes,
    sheet_name: str | None = None,
) -> tuple[str, list[list[object]], tuple[str, ...], str | None]:
    extension = Path(filename).suffix.lower().lstrip(".")
    if extension == "csv":
        return extension, _read_csv_rows(content), (), None
    if extension == "xlsx":
        rows, sheets, selected = _read_xlsx_rows(content, sheet_name)
        return extension, rows, sheets, selected
    raise ImportFileError("Mapeamento manual esta disponivel apenas para CSV e XLSX.")


def inspect_import_file(
    filename: str,
    content: bytes,
    *,
    sheet_name: str | None = None,
) -> ImportInspection:
    if not content:
        raise ImportFileError("O arquivo esta vazio.")

    extension = Path(filename).suffix.lower().lstrip(".")
    if extension == "ofx":
        blocks = _ofx_blocks(content)
        preview_rows: list[tuple[str, ...]] = [
            ("Data", "Descricao", "Valor", "Identificador")
        ]
        for block in blocks[: MAX_INSPECTION_ROWS - 1]:
            preview_rows.append(
                (
                    _ofx_value(block, "DTPOSTED"),
                    _ofx_value(block, "NAME") or _ofx_value(block, "MEMO"),
                    _ofx_value(block, "TRNAMT"),
                    _ofx_value(block, "FITID"),
                )
            )
        return ImportInspection(
            source_type="ofx",
            sheets=(),
            selected_sheet=None,
            total_rows=len(blocks),
            max_columns=4,
            rows=tuple(preview_rows),
            mapping_required=False,
        )

    source_type, raw_rows, sheets, selected = _read_tabular_file(
        filename,
        content,
        sheet_name,
    )
    if not raw_rows:
        raise ImportFileError("Nenhuma linha foi encontrada no arquivo.")

    max_columns = max((len(row) for row in raw_rows), default=0)
    preview = tuple(
        tuple(_clean_text(value) for value in row)
        for row in raw_rows[:MAX_INSPECTION_ROWS]
    )
    return ImportInspection(
        source_type=source_type,
        sheets=sheets,
        selected_sheet=selected,
        total_rows=len(raw_rows),
        max_columns=max_columns,
        rows=preview,
        mapping_required=True,
    )


def _parse_date(value: object, date_format: str) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = _clean_text(value)
    if not text:
        raise ValueError("Data ausente.")

    patterns_by_mode = {
        "dmy": ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"),
        "mdy": ("%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y"),
        "ymd": ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"),
        "auto": (
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%Y/%m/%d",
            "%d/%m/%y",
            "%m/%d/%Y",
            "%Y%m%d",
        ),
    }
    patterns = patterns_by_mode.get(date_format)
    if patterns is None:
        raise ValueError("Formato de data nao suportado.")

    if date_format == "auto":
        iso_candidate = text.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(iso_candidate)
            return parsed.replace(tzinfo=None)
        except ValueError:
            pass

    for pattern in patterns:
        candidate = text[:8] if pattern == "%Y%m%d" else text[:10]
        try:
            return datetime.strptime(candidate, pattern)
        except ValueError:
            continue

    raise ValueError(f"Data invalida: {text}")


def _parse_signed_amount(value: object, decimal_separator: str) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return Decimal(str(value))

    text = _clean_text(value)
    if not text:
        raise ValueError("Valor ausente.")

    negative_parentheses = text.startswith("(") and text.endswith(")")
    negative_trailing = text.endswith("-")
    text = text.replace("R$", "").replace("US$", "").replace("$", "")
    text = text.replace("\u20ac", "").replace("\u00a3", "").replace(" ", "")
    text = text.strip("()")
    if negative_trailing:
        text = text[:-1]

    if decimal_separator == "comma":
        text = text.replace(".", "").replace(",", ".")
    elif decimal_separator == "dot":
        text = text.replace(",", "")
    elif decimal_separator == "auto":
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")
    else:
        raise ValueError("Separador decimal nao suportado.")

    text = re.sub(r"[^0-9+\-.]", "", text)
    try:
        amount = Decimal(text)
    except InvalidOperation as error:
        raise ValueError(f"Valor invalido: {value}") from error

    if negative_parentheses or negative_trailing:
        amount = -abs(amount)
    return amount


def _expense_amount(value: object, decimal_separator: str, amount_mode: str) -> Decimal:
    signed = _parse_signed_amount(value, decimal_separator)
    if signed == 0:
        raise ValueError("O valor deve ser diferente de zero.")

    if amount_mode == "positive" and signed < 0:
        raise IgnoredImportRow("Valor negativo ignorado pela configuracao.")
    if amount_mode == "negative" and signed > 0:
        raise IgnoredImportRow("Valor positivo ignorado pela configuracao.")
    if amount_mode not in {"all", "positive", "negative"}:
        raise ValueError("Regra de sinal nao suportada.")

    return abs(signed).quantize(Decimal("0.01"))


def _fingerprint(
    source_type: str,
    transaction_date: datetime,
    place: str,
    amount: Decimal,
    external_id: str | None,
) -> str:
    raw = "|".join(
        (
            source_type,
            transaction_date.date().isoformat(),
            f"{amount:.2f}",
            _normalize(place),
            _normalize(external_id or ""),
        )
    )
    return sha256(raw.encode("utf-8")).hexdigest()


def _cell(row: list[object], column: int | None) -> object | None:
    if column is None or column < 0 or column >= len(row):
        return None
    return row[column]


def _validate_mapping(mapping: ImportColumnMapping, max_columns: int) -> None:
    if mapping.data_start_row < 1:
        raise ImportFileError("A primeira linha de dados deve ser maior que zero.")
    if not mapping.description_columns:
        raise ImportFileError("Selecione ao menos uma coluna de descricao.")

    required_columns = (
        mapping.date_column,
        mapping.amount_column,
        *mapping.description_columns,
    )
    optional_columns = (
        (mapping.external_id_column,)
        if mapping.external_id_column is not None
        else ()
    )
    for column in (*required_columns, *optional_columns):
        if column < 0 or column >= max_columns:
            raise ImportFileError(
                f"A coluna {column + 1} nao existe no arquivo selecionado."
            )


def _row_from_values(
    *,
    row_number: int,
    values: list[object],
    source_type: str,
    mapping: ImportColumnMapping,
) -> ParsedImportRow:
    place = " - ".join(
        text
        for text in (
            _clean_text(_cell(values, column))
            for column in mapping.description_columns
        )
        if text
    )
    external_value = _cell(values, mapping.external_id_column)
    external_id = _clean_text(external_value) or None

    try:
        transaction_date = _parse_date(
            _cell(values, mapping.date_column),
            mapping.date_format,
        )
        if not place:
            raise ValueError("Descricao ou estabelecimento ausente.")
        amount = _expense_amount(
            _cell(values, mapping.amount_column),
            mapping.decimal_separator,
            mapping.amount_mode,
        )
        fingerprint = _fingerprint(
            source_type,
            transaction_date,
            place,
            amount,
            external_id,
        )
        return ParsedImportRow(
            row_number=row_number,
            purchase_date=transaction_date,
            purchase_place=place[:255],
            purchase_value=amount,
            external_id=external_id,
            fingerprint=fingerprint,
        )
    except IgnoredImportRow as error:
        return ParsedImportRow(
            row_number=row_number,
            purchase_date=None,
            purchase_place=place[:255] or None,
            purchase_value=None,
            external_id=external_id,
            fingerprint=None,
            classification="ignored",
            error_message=str(error),
        )
    except ValueError as error:
        return ParsedImportRow(
            row_number=row_number,
            purchase_date=None,
            purchase_place=place[:255] or None,
            purchase_value=None,
            external_id=external_id,
            fingerprint=None,
            classification="invalid",
            error_message=str(error),
        )


def _parse_tabular(
    filename: str,
    content: bytes,
    mapping: ImportColumnMapping,
) -> ParsedImportFile:
    source_type, raw_rows, _, _ = _read_tabular_file(
        filename,
        content,
        mapping.sheet_name,
    )
    if not raw_rows:
        raise ImportFileError("Nenhuma linha foi encontrada no arquivo.")

    max_columns = max((len(row) for row in raw_rows), default=0)
    _validate_mapping(mapping, max_columns)

    parsed_rows: list[ParsedImportRow] = []
    for row_number, values in enumerate(raw_rows, start=1):
        if row_number < mapping.data_start_row:
            continue
        if not any(_clean_text(value) for value in values):
            continue
        parsed_rows.append(
            _row_from_values(
                row_number=row_number,
                values=values,
                source_type=source_type,
                mapping=mapping,
            )
        )

    return ParsedImportFile(source_type=source_type, rows=tuple(parsed_rows))


def _parse_ofx(content: bytes) -> ParsedImportFile:
    rows: list[ParsedImportRow] = []
    for index, block in enumerate(_ofx_blocks(content), start=1):
        mapping = ImportColumnMapping(
            data_start_row=1,
            date_column=0,
            description_columns=(1,),
            amount_column=2,
            external_id_column=3,
            date_format="ymd",
            decimal_separator="dot",
            amount_mode="all",
        )
        values: list[object] = [
            _ofx_value(block, "DTPOSTED"),
            _ofx_value(block, "NAME") or _ofx_value(block, "MEMO"),
            _ofx_value(block, "TRNAMT"),
            _ofx_value(block, "FITID"),
        ]
        rows.append(
            _row_from_values(
                row_number=index,
                values=values,
                source_type="ofx",
                mapping=mapping,
            )
        )
    return ParsedImportFile(source_type="ofx", rows=tuple(rows))


def parse_import_file(
    filename: str,
    content: bytes,
    mapping: ImportColumnMapping | None = None,
) -> ParsedImportFile:
    if not content:
        raise ImportFileError("O arquivo esta vazio.")

    extension = Path(filename).suffix.lower().lstrip(".")
    if extension == "ofx":
        parsed = _parse_ofx(content)
    elif extension in {"csv", "xlsx"}:
        if mapping is None:
            raise ImportFileError(
                "Defina o mapeamento das colunas antes de analisar CSV ou XLSX."
            )
        parsed = _parse_tabular(filename, content, mapping)
    else:
        raise ImportFileError("Formato nao suportado. Use CSV, XLSX ou OFX.")

    if not parsed.rows:
        raise ImportFileError("Nenhuma linha de transacao foi encontrada.")
    if len(parsed.rows) > MAX_IMPORT_ROWS:
        raise ImportFileError(
            f"O arquivo excede o limite de {MAX_IMPORT_ROWS} transacoes."
        )
    return parsed
