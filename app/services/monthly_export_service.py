from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

from app.repositories.receivable_repository import (
    ReceivableRepository,
)
from app.repositories.report_repository import (
    ReportRepository,
)


class MonthlyExportService:
    CENT = Decimal("0.01")
    CURRENCY_FORMAT = 'R$ #,##0.00'
    DATE_FORMAT = 'dd/mm/yyyy'
    HEADER_FILL = '1F6B45'
    HEADER_FONT = 'FFFFFF'
    ALT_FILL = 'EAF4EE'
    BORDER_COLOR = 'B7C9BD'

    def __init__(
        self,
        *,
        report_repository: ReportRepository,
        receivable_repository: ReceivableRepository,
    ):
        self.report_repository = report_repository
        self.receivable_repository = receivable_repository

    def build(
        self,
        *,
        year: int,
        month: int,
    ) -> tuple[bytes, str]:
        self._validate_period(year=year, month=month)

        expenses = (
            self.report_repository
            .list_purchases_for_month(
                year=year,
                month=month,
            )
        )
        debt_summary = (
            self.receivable_repository
            .list_open_summary_for_month(
                year=year,
                month=month,
            )
        )
        debt_details = (
            self.receivable_repository
            .list_open_details_for_month(
                year=year,
                month=month,
            )
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatorio mensal"

        expense_headers = [
            "Valor",
            "Data",
            "Obs",
            "Parcelado",
            "Nr parcelas",
            "Dividido",
            "Pessoas divididas",
        ]
        debt_headers = [
            "Quem deve?",
            "Quanto deve",
        ]

        for column, value in enumerate(
            expense_headers,
            start=1,
        ):
            sheet.cell(
                row=1,
                column=column,
                value=value,
            )

        for column, value in enumerate(
            debt_headers,
            start=9,
        ):
            sheet.cell(
                row=1,
                column=column,
                value=value,
            )

        for row_index, expense in enumerate(
            expenses,
            start=2,
        ):
            people = ", ".join(
                relation.person.name
                for relation in expense.people
            )
            installment_count = max(
                (
                    item.total_installments
                    for item in expense.installments
                ),
                default=(
                    len(expense.installments)
                    if expense.is_installment
                    else 1
                ),
            )
            observation = expense.purchase_place
            if expense.notes:
                observation = (
                    f"{expense.purchase_place} - "
                    f"{expense.notes}"
                )

            values = [
                self._money(expense.purchase_value),
                expense.purchase_date.date(),
                observation,
                "Sim" if expense.is_installment else "Nao",
                installment_count,
                "Sim" if expense.is_shared else "Nao",
                people,
            ]

            for column, value in enumerate(
                values,
                start=1,
            ):
                sheet.cell(
                    row=row_index,
                    column=column,
                    value=value,
                )

        for row_index, row in enumerate(
            debt_summary,
            start=2,
        ):
            sheet.cell(
                row=row_index,
                column=9,
                value=row.person_name,
            )
            sheet.cell(
                row=row_index,
                column=10,
                value=self._money(row.total),
            )

        self._format_main_sheet(
            sheet,
            expense_rows=len(expenses),
            debt_rows=len(debt_summary),
        )
        self._add_detail_sheet(
            workbook,
            debt_details,
        )

        output = BytesIO()
        workbook.save(output)
        filename = (
            f"financebot_relatorio_"
            f"{year}_{month:02d}.xlsx"
        )
        return output.getvalue(), filename

    def _format_main_sheet(
        self,
        sheet,
        *,
        expense_rows: int,
        debt_rows: int,
    ) -> None:
        thin = Side(
            style="thin",
            color=self.BORDER_COLOR,
        )
        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )
        header_fill = PatternFill(
            "solid",
            fgColor=self.HEADER_FILL,
        )
        header_font = Font(
            bold=True,
            color=self.HEADER_FONT,
        )

        for cell in sheet[1]:
            if cell.column == 8:
                continue
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = border

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:G{max(expense_rows + 1, 2)}"
        )

        for row in range(2, expense_rows + 2):
            for column in range(1, 8):
                cell = sheet.cell(row=row, column=column)
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
            sheet.cell(row=row, column=1).number_format = self.CURRENCY_FORMAT
            sheet.cell(row=row, column=2).number_format = self.DATE_FORMAT

        for row in range(2, debt_rows + 2):
            for column in range(9, 11):
                cell = sheet.cell(row=row, column=column)
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
            sheet.cell(row=row, column=10).number_format = self.CURRENCY_FORMAT

        if expense_rows:
            expense_table = Table(
                displayName="DespesasMensais",
                ref=f"A1:G{expense_rows + 1}",
            )
            expense_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(expense_table)

        if debt_rows:
            debt_table = Table(
                displayName="ResumoQuemDeve",
                ref=f"I1:J{debt_rows + 1}",
            )
            debt_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(debt_table)

        widths = {
            "A": 15,
            "B": 13,
            "C": 38,
            "D": 13,
            "E": 13,
            "F": 12,
            "G": 32,
            "H": 3,
            "I": 22,
            "J": 16,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[1].height = 24

    def _add_detail_sheet(
        self,
        workbook: Workbook,
        rows,
    ) -> None:
        sheet = workbook.create_sheet(
            "Detalhes por pessoa"
        )
        headers = [
            "Pessoa",
            "Quanto deve",
            "Data",
            "Estabelecimento",
            "Observacao",
        ]
        sheet.append(headers)

        for row in rows:
            sheet.append(
                [
                    row.person_name,
                    self._money(row.amount),
                    row.purchase_date.date(),
                    row.purchase_place,
                    row.notes or "",
                ]
            )

        header_fill = PatternFill(
            "solid",
            fgColor=self.HEADER_FILL,
        )
        header_font = Font(
            bold=True,
            color=self.HEADER_FONT,
        )
        thin = Side(
            style="thin",
            color=self.BORDER_COLOR,
        )
        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
            )

        for row_index in range(2, len(rows) + 2):
            for column in range(1, 6):
                cell = sheet.cell(
                    row=row_index,
                    column=column,
                )
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
            sheet.cell(
                row=row_index,
                column=2,
            ).number_format = self.CURRENCY_FORMAT
            sheet.cell(
                row=row_index,
                column=3,
            ).number_format = self.DATE_FORMAT

        if rows:
            table = Table(
                displayName="DetalhesRecebiveis",
                ref=f"A1:E{len(rows) + 1}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

        sheet.freeze_panes = "A2"
        widths = {
            "A": 22,
            "B": 16,
            "C": 13,
            "D": 28,
            "E": 38,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

    @classmethod
    def _money(
        cls,
        value,
    ) -> Decimal:
        return Decimal(
            str(value)
        ).quantize(cls.CENT)

    @staticmethod
    def _validate_period(
        *,
        year: int,
        month: int,
    ) -> None:
        if year < 2000 or year > 2100:
            raise ValueError("Ano invalido.")
        if month < 1 or month > 12:
            raise ValueError("Mes invalido.")
